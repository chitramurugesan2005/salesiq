from flask import Blueprint, request, jsonify, Response
from models import db, Sale, Product, Customer
from sqlalchemy import func
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import pandas as pd
import openpyxl
import io
from ai_ml.forecast import run_forecast
from ai_ml.segment import run_segmentation

reports_bp = Blueprint('reports', __name__)

# ── Helper: get date range ──
def get_date_range(period):
    today = datetime.today().date()
    if period == 'month':
        start = today - timedelta(days=30)
    elif period == 'quarter':
        start = today - timedelta(days=90)
    elif period == 'year':
        start = today - timedelta(days=365)
    else:
        start = today - timedelta(days=90)
    return start, today


# ── Helper: get summary data ──
def get_summary_data(start, end):
    summary = db.session.query(
        func.sum(Sale.revenue),
        func.count(Sale.id),
        func.sum(Sale.units)
    ).filter(
        Sale.sale_date >= start,
        Sale.sale_date <= end
    ).first()

    top_products = db.session.query(
        Product.name,
        Product.category,
        func.sum(Sale.revenue).label('revenue'),
        func.sum(Sale.units).label('units')
    ).join(
        Sale, Sale.product_id == Product.id
    ).filter(
        Sale.sale_date >= start,
        Sale.sale_date <= end
    ).group_by(
        Product.id, Product.name, Product.category
    ).order_by(
        func.sum(Sale.revenue).desc()
    ).limit(5).all()

    regions = db.session.query(
        Sale.region,
        func.sum(Sale.revenue).label('revenue'),
        func.count(Sale.id).label('orders')
    ).filter(
        Sale.sale_date >= start,
        Sale.sale_date <= end
    ).group_by(Sale.region).order_by(
        func.sum(Sale.revenue).desc()
    ).all()

    return summary, top_products, regions


# ── Helper: get customer data ──
def get_customer_data():
    customers = Customer.query.order_by(
        Customer.total_spend.desc()
    ).limit(10).all()
    return customers


# ── Helper: get category data ──
def get_category_data(start, end):
    categories = db.session.query(
        Product.category,
        func.sum(Sale.revenue).label('revenue'),
        func.sum(Sale.units).label('units'),
        func.count(Sale.id).label('orders')
    ).join(
        Sale, Sale.product_id == Product.id
    ).filter(
        Sale.sale_date >= start,
        Sale.sale_date <= end
    ).group_by(Product.category).order_by(
        func.sum(Sale.revenue).desc()
    ).all()
    return categories


# ── PDF table style helper ──
def make_table_style(header_color):
    return TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0), colors.HexColor(header_color)),
        ('TEXTCOLOR',      (0, 0), (-1, 0), colors.white),
        ('FONTNAME',       (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, 0), 11),
        ('ALIGN',          (0, 0), (-1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
            [colors.HexColor('#f5f5f5'), colors.white]),
        ('GRID',           (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING',        (0, 0), (-1, -1), 8),
    ])


# ══════════════════════════════════════════
# ── Generate PDF report ──
# ══════════════════════════════════════════
@reports_bp.route('/pdf', methods=['GET'])
def generate_pdf():
    period      = request.args.get('period', 'quarter')
    title       = request.args.get('title', 'SalesIQ Report')
    prepared_by = request.args.get('prepared_by', 'Admin')
    report_type = request.args.get('report_type', 'full')
    start, end  = get_date_range(period)

    summary, top_products, regions = get_summary_data(start, end)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story  = []

    # ── Title ──
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f'Period: {start} to {end} | Prepared by: {prepared_by} | Type: {report_type.upper()}',
        styles['Normal']
    ))
    story.append(Spacer(1, 20))

    # ══ FULL REPORT ══
    if report_type == 'full':
        # KPI Summary
        story.append(Paragraph('KPI Summary', styles['Heading1']))
        story.append(Spacer(1, 8))
        kpi_data = [
            ['Metric', 'Value'],
            ['Total Revenue', f"${float(summary[0] or 0):,.2f}"],
            ['Total Orders',  str(int(summary[1] or 0))],
            ['Total Units',   str(int(summary[2] or 0))],
        ]
        kpi_table = Table(kpi_data, colWidths=[250, 200])
        kpi_table.setStyle(make_table_style('#4f8ef7'))
        story.append(kpi_table)
        story.append(Spacer(1, 20))

        # Top Products
        story.append(Paragraph('Top 5 Products', styles['Heading1']))
        story.append(Spacer(1, 8))
        prod_data = [['Rank', 'Product', 'Category', 'Revenue', 'Units']]
        for i, row in enumerate(top_products):
            prod_data.append([str(i+1), row[0], row[1],
                f"${float(row[2] or 0):,.2f}", str(int(row[3] or 0))])
        prod_table = Table(prod_data, colWidths=[40, 160, 100, 100, 60])
        prod_table.setStyle(make_table_style('#3ecf8e'))
        story.append(prod_table)
        story.append(Spacer(1, 20))

        # Regional Breakdown
        story.append(Paragraph('Regional Breakdown', styles['Heading1']))
        story.append(Spacer(1, 8))
        reg_data = [['Region', 'Revenue', 'Orders']]
        for row in regions:
            reg_data.append([row[0], f"${float(row[1] or 0):,.2f}", str(int(row[2] or 0))])
        reg_table = Table(reg_data, colWidths=[160, 160, 140])
        reg_table.setStyle(make_table_style('#f0a04b'))
        story.append(reg_table)
        story.append(Spacer(1, 20))

        # Customer Summary
        customers = get_customer_data()
        story.append(Paragraph('Top Customers', styles['Heading1']))
        story.append(Spacer(1, 8))
        cust_data = [['Name', 'Segment', 'Orders', 'Total Spend', 'LTV']]
        for c in customers[:5]:
            cust_data.append([
                c.name, c.segment, str(c.total_orders),
                f"${float(c.total_spend):,.2f}",
                f"${float(c.ltv):,.2f}"
            ])
        cust_table = Table(cust_data, colWidths=[130, 80, 60, 100, 90])
        cust_table.setStyle(make_table_style('#a78bfa'))
        story.append(cust_table)
        story.append(Spacer(1, 20))

        # AI Forecast
        story.append(Paragraph('AI Sales Forecast', styles['Heading1']))
        story.append(Spacer(1, 8))
        try:
            forecast_result, err = run_forecast()
            if not err and forecast_result:
                fc_data = [['Month', 'Forecasted Revenue']]
                for f in forecast_result.get('forecast', [])[:6]:
                    fc_data.append([
                        str(f.get('label', '')),
                        f"${float(f.get('value', 0)):,.2f}"
                    ])
                fc_data.append([
                    'Model R² Score',
                    str(round(forecast_result.get('r_squared', 0), 3))
                ])
                fc_table = Table(fc_data, colWidths=[200, 200])
                fc_table.setStyle(make_table_style('#8b5cf6'))
                story.append(fc_table)
            else:
                story.append(Paragraph('Forecast data unavailable.', styles['Normal']))
        except Exception as e:
            story.append(Paragraph(f'Forecast error: {str(e)}', styles['Normal']))

        story.append(Spacer(1, 20))

        # AI Segmentation
        story.append(Paragraph('Customer Segmentation — K-Means Clusters', styles['Heading1']))
        story.append(Spacer(1, 8))
        try:
            seg_result, err = run_segmentation()
            if not err and seg_result:
                seg_summary = seg_result.get('summary', {})
                k           = seg_result.get('k', 3)
                seg_data    = [['Segment', 'Customer Count']]
                for segment, count in seg_summary.items():
                    seg_data.append([segment, str(count)])
                seg_data.append(['Total Clusters (K)', str(k)])
                seg_table = Table(seg_data, colWidths=[250, 200])
                seg_table.setStyle(make_table_style('#a78bfa'))
                story.append(seg_table)
            else:
                story.append(Paragraph('Segmentation data unavailable.', styles['Normal']))
        except Exception as e:
            story.append(Paragraph(f'Segmentation error: {str(e)}', styles['Normal']))

    # ══ SALES REPORT ══
    elif report_type == 'sales':
        story.append(Paragraph('Sales KPI Summary', styles['Heading1']))
        story.append(Spacer(1, 8))
        kpi_data = [
            ['Metric', 'Value'],
            ['Total Revenue',    f"${float(summary[0] or 0):,.2f}"],
            ['Total Orders',     str(int(summary[1] or 0))],
            ['Total Units Sold', str(int(summary[2] or 0))],
            ['Avg Order Value',  f"${float(summary[0] or 0) / max(int(summary[1] or 1), 1):,.2f}"],
        ]
        kpi_table = Table(kpi_data, colWidths=[250, 200])
        kpi_table.setStyle(make_table_style('#22c55e'))
        story.append(kpi_table)
        story.append(Spacer(1, 20))

        story.append(Paragraph('Top 5 Products by Revenue', styles['Heading1']))
        story.append(Spacer(1, 8))
        prod_data = [['Rank', 'Product', 'Category', 'Revenue', 'Units']]
        for i, row in enumerate(top_products):
            prod_data.append([str(i+1), row[0], row[1],
                f"${float(row[2] or 0):,.2f}", str(int(row[3] or 0))])
        prod_table = Table(prod_data, colWidths=[40, 160, 100, 100, 60])
        prod_table.setStyle(make_table_style('#3ecf8e'))
        story.append(prod_table)
        story.append(Spacer(1, 20))

        categories = get_category_data(start, end)
        story.append(Paragraph('Revenue by Category', styles['Heading1']))
        story.append(Spacer(1, 8))
        cat_data  = [['Category', 'Revenue', 'Units', 'Orders']]
        total_rev = float(summary[0] or 0)
        for row in categories:
            pct = (float(row[1] or 0) / total_rev * 100) if total_rev > 0 else 0
            cat_data.append([
                row[0],
                f"${float(row[1] or 0):,.2f} ({pct:.1f}%)",
                str(int(row[2] or 0)),
                str(int(row[3] or 0))
            ])
        cat_table = Table(cat_data, colWidths=[140, 160, 80, 80])
        cat_table.setStyle(make_table_style('#4f8ef7'))
        story.append(cat_table)
        story.append(Spacer(1, 20))

        story.append(Paragraph('Regional Sales Performance', styles['Heading1']))
        story.append(Spacer(1, 8))
        reg_data = [['Region', 'Revenue', 'Orders', 'Avg Order Value']]
        for row in regions:
            avg = float(row[1] or 0) / max(int(row[2] or 1), 1)
            reg_data.append([
                row[0],
                f"${float(row[1] or 0):,.2f}",
                str(int(row[2] or 0)),
                f"${avg:,.2f}"
            ])
        reg_table = Table(reg_data, colWidths=[120, 130, 80, 130])
        reg_table.setStyle(make_table_style('#f0a04b'))
        story.append(reg_table)

    # ══ CUSTOMER REPORT ══
    elif report_type == 'customer':
        customers = get_customer_data()

        story.append(Paragraph('Customer Segment Summary', styles['Heading1']))
        story.append(Spacer(1, 8))
        seg_counts = {}
        seg_spend  = {}
        for c in Customer.query.all():
            seg = c.segment
            seg_counts[seg] = seg_counts.get(seg, 0) + 1
            seg_spend[seg]  = seg_spend.get(seg, 0) + float(c.total_spend)
        seg_data = [['Segment', 'Customers', 'Total Spend', 'Avg Spend']]
        for seg in seg_counts:
            count = seg_counts[seg]
            spend = seg_spend[seg]
            seg_data.append([seg, str(count),
                f"${spend:,.2f}", f"${spend/max(count,1):,.2f}"])
        seg_table = Table(seg_data, colWidths=[130, 80, 130, 120])
        seg_table.setStyle(make_table_style('#a78bfa'))
        story.append(seg_table)
        story.append(Spacer(1, 20))

        story.append(Paragraph('Top 10 Customers by Lifetime Value', styles['Heading1']))
        story.append(Spacer(1, 8))
        cust_data = [['#', 'Name', 'Email', 'Segment', 'Orders', 'Total Spend', 'LTV']]
        for i, c in enumerate(customers):
            cust_data.append([
                str(i+1), c.name, c.email, c.segment,
                str(c.total_orders),
                f"${float(c.total_spend):,.2f}",
                f"${float(c.ltv):,.2f}"
            ])
        cust_table = Table(cust_data, colWidths=[25, 90, 120, 70, 45, 80, 60])
        cust_table.setStyle(make_table_style('#0fb8a0'))
        story.append(cust_table)
        story.append(Spacer(1, 20))

        story.append(Paragraph('Customer Distribution by Region', styles['Heading1']))
        story.append(Spacer(1, 8))
        reg_dist = {}
        for c in Customer.query.all():
            reg_dist[c.region] = reg_dist.get(c.region, 0) + 1
        reg_data = [['Region', 'Customers']]
        for reg, count in sorted(reg_dist.items(), key=lambda x: -x[1]):
            reg_data.append([reg, str(count)])
        reg_table = Table(reg_data, colWidths=[200, 200])
        reg_table.setStyle(make_table_style('#3b82f6'))
        story.append(reg_table)
        story.append(Spacer(1, 20))

        # AI Segmentation for customer report
        story.append(Paragraph('K-Means Cluster Analysis', styles['Heading1']))
        story.append(Spacer(1, 8))
        try:
            seg_result, err = run_segmentation()
            if not err and seg_result:
                seg_summary = seg_result.get('summary', {})
                k           = seg_result.get('k', 3)
                seg_data2   = [['Segment', 'Customer Count']]
                for segment, count in seg_summary.items():
                    seg_data2.append([segment, str(count)])
                seg_data2.append(['Total Clusters (K)', str(k)])
                seg_table2 = Table(seg_data2, colWidths=[250, 200])
                seg_table2.setStyle(make_table_style('#a78bfa'))
                story.append(seg_table2)
            else:
                story.append(Paragraph('Segmentation data unavailable.', styles['Normal']))
        except Exception as e:
            story.append(Paragraph(f'Segmentation error: {str(e)}', styles['Normal']))

    # ── Footer ──
    story.append(Spacer(1, 30))
    story.append(Paragraph(
        f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")} | SalesIQ Analytics Platform',
        styles['Normal']
    ))

    doc.build(story)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition':
                f'attachment; filename=salesiq_{report_type}_report.pdf'
        }
    )


# ══════════════════════════════════════════
# ── Generate Excel report ──
# ══════════════════════════════════════════
@reports_bp.route('/excel', methods=['GET'])
def generate_excel():
    period      = request.args.get('period', 'quarter')
    report_type = request.args.get('report_type', 'full')
    start, end  = get_date_range(period)

    summary, top_products, regions = get_summary_data(start, end)

    sales = Sale.query.join(Product).filter(
        Sale.sale_date >= start,
        Sale.sale_date <= end
    ).order_by(Sale.sale_date.desc()).all()

    buffer = io.BytesIO()
    wb     = openpyxl.Workbook()

    ws1       = wb.active
    ws1.title = 'KPI Summary'
    ws1.append(['SalesIQ Report'])
    ws1.append([f'Period: {start} to {end}'])
    ws1.append([f'Report Type: {report_type.upper()}'])
    ws1.append([])
    ws1.append(['Metric', 'Value'])
    ws1.append(['Total Revenue',    float(summary[0] or 0)])
    ws1.append(['Total Orders',     int(summary[1] or 0)])
    ws1.append(['Total Units Sold', int(summary[2] or 0)])
    ws1.append(['Avg Order Value',
        float(summary[0] or 0) / max(int(summary[1] or 1), 1)])

    if report_type in ('full', 'sales'):
        ws2 = wb.create_sheet('Top Products')
        ws2.append(['Rank', 'Product', 'Category', 'Revenue', 'Units'])
        for i, row in enumerate(top_products):
            ws2.append([i+1, row[0], row[1],
                float(row[2] or 0), int(row[3] or 0)])

        categories = get_category_data(start, end)
        ws3 = wb.create_sheet('By Category')
        ws3.append(['Category', 'Revenue', 'Units', 'Orders'])
        for row in categories:
            ws3.append([row[0], float(row[1] or 0),
                int(row[2] or 0), int(row[3] or 0)])

        ws4 = wb.create_sheet('Regional')
        ws4.append(['Region', 'Revenue', 'Orders'])
        for row in regions:
            ws4.append([row[0], float(row[1] or 0), int(row[2] or 0)])

    if report_type in ('full', 'customer'):
        customers = get_customer_data()
        ws5 = wb.create_sheet('Top Customers')
        ws5.append(['#', 'Name', 'Email', 'Region',
            'Segment', 'Orders', 'Total Spend', 'LTV'])
        for i, c in enumerate(customers):
            ws5.append([i+1, c.name, c.email, c.region,
                c.segment, c.total_orders,
                float(c.total_spend), float(c.ltv)])

        ws6 = wb.create_sheet('Segment Summary')
        ws6.append(['Segment', 'Count', 'Total Spend', 'Avg Spend'])
        seg_counts = {}
        seg_spend  = {}
        for c in Customer.query.all():
            seg_counts[c.segment] = seg_counts.get(c.segment, 0) + 1
            seg_spend[c.segment]  = seg_spend.get(c.segment, 0) + float(c.total_spend)
        for seg in seg_counts:
            ws6.append([seg, seg_counts[seg], seg_spend[seg],
                seg_spend[seg] / max(seg_counts[seg], 1)])

    if report_type == 'full':
        ws7 = wb.create_sheet('Raw Sales Data')
        ws7.append(['ID', 'Product', 'Category', 'Date',
            'Region', 'Units', 'Price', 'Revenue', 'Status', 'Notes'])
        for s in sales:
            ws7.append([
                s.id, s.product.name, s.product.category,
                s.sale_date.strftime('%Y-%m-%d'),
                s.region, s.units, float(s.product.price),
                float(s.revenue), s.status, s.notes
            ])

        # AI Forecast sheet
        try:
            forecast_result, err = run_forecast()
            if not err and forecast_result:
                ws8 = wb.create_sheet('AI Forecast')
                ws8.append(['Month', 'Forecasted Revenue'])
                for f in forecast_result.get('forecast', []):
                    ws8.append([
                        str(f.get('label', '')),
                        float(f.get('value', 0))
                    ])
                ws8.append(['R² Score', forecast_result.get('r_squared', 0)])
        except Exception:
            pass

    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition':
                f'attachment; filename=salesiq_{report_type}_report.xlsx'
        }
    )


# ── Generate CSV report ──
@reports_bp.route('/csv', methods=['GET'])
def generate_csv():
    period      = request.args.get('period', 'quarter')
    report_type = request.args.get('report_type', 'full')
    start, end  = get_date_range(period)

    if report_type == 'customer':
        customers = get_customer_data()
        data = [{
            'Name'        : c.name,
            'Email'       : c.email,
            'Region'      : c.region,
            'Segment'     : c.segment,
            'Orders'      : c.total_orders,
            'Total Spend' : float(c.total_spend),
            'LTV'         : float(c.ltv)
        } for c in customers]
    else:
        sales = Sale.query.join(Product).filter(
            Sale.sale_date >= start,
            Sale.sale_date <= end
        ).order_by(Sale.sale_date.desc()).all()
        data = [{
            'ID'       : s.id,
            'Product'  : s.product.name,
            'Category' : s.product.category,
            'Date'     : s.sale_date.strftime('%Y-%m-%d'),
            'Region'   : s.region,
            'Units'    : s.units,
            'Price'    : float(s.product.price),
            'Revenue'  : float(s.revenue),
            'Status'   : s.status,
            'Notes'    : s.notes
        } for s in sales]

    df     = pd.DataFrame(data)
    output = io.StringIO()
    df.to_csv(output, index=False)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition':
                f'attachment; filename=salesiq_{report_type}_report.csv'
        }
    )


# ── Report history ──
@reports_bp.route('/history', methods=['GET'])
def get_history():
    history = [
        {'name': 'Q2 2026 Full Report',  'format': 'PDF',   'date': 'Today',      'status': 'Ready'},
        {'name': 'Sales Performance',     'format': 'Excel', 'date': 'Yesterday',  'status': 'Ready'},
        {'name': 'Customer Analytics',    'format': 'PDF',   'date': '2 days ago', 'status': 'Ready'},
    ]
    return jsonify(history), 200